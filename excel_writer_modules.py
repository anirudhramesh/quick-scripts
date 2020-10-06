import pandas as pd
from os.path import exists
from openpyxl import load_workbook


def write_dataframe_to_new_file(filename, dataframe, sheetname='Sheet1', set_autofilter: bool = False, startrow=0,
                                startcol=0, index=False, adjust_col_width: bool = True, money_fmt: list = None,
                                perc_fmt_no_decimal: list = None, perc_fmt_decimal: list = None,
                                decimal_fmt: list = None):

    for fmt in [money_fmt, perc_fmt_no_decimal, perc_fmt_decimal, decimal_fmt]:
        if fmt is None:
            fmt = []
    columns_to_format = {'money_fmt': money_fmt, 'perc_fmt_no_decimal': perc_fmt_no_decimal,
                         'perc_fmt_decimal': perc_fmt_decimal, 'decimal_fmt': decimal_fmt}

    writer = pd.ExcelWriter(filename)
    write_frame = dataframe

    write_frame.to_excel(writer, index=index)
    book = writer.book
    sheet = writer.sheets[sheetname]

    column_format_objects = {
        'money_fmt': book.add_format({'num_format': '$#,##0.00'}),
        'perc_fmt_no_decimal': book.add_format({'num_format': '0%'}),
        'perc_fmt_decimal': book.add_format({'num_format': '0.00%'}),
        'decimal_fmt': book.add_format({'num_format': '0.00'})
    }

    col_offset = startcol + (1 if index else 0)
    for fmt in column_format_objects:
        for col in [dataframe.columns.get_loc(c)+col_offset for c in columns_to_format.get(fmt, [])]:
            sheet.set_column(col, col, cell_format=column_format_objects[fmt])

    if adjust_col_width:
        for ind in range(len(write_frame.columns)):
            max_len = max(len(write_frame.columns[ind]), write_frame.iloc[:, ind].astype(str).map(len).max()) + 1
            sheet.set_column(ind, ind, max_len)

    if set_autofilter:
        sheet.autofilter(0, 0, write_frame.shape[0] - 1, write_frame.shape[1] - 1)

    writer.save(), writer.close()


def rewrite_existing_file(filename, dataframe, existing_data_dimensions=((1, 20), (1, 4000)), sheet_name='Sheet1',
                          adjust_col_width: bool = True, columns_to_format=None):
    if columns_to_format is None:
        columns_to_format = {'money_fmt': [], 'perc_fmt_no_decimal': [],
                             'perc_fmt_decimal': [], 'decimal_fmt': []}

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    if exists(filename):
        book = load_workbook(filename)
        writer.book = book
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        sheet = book[sheet_name]
        for col in range(*existing_data_dimensions[0]):
            for row in range(*existing_data_dimensions[1]):
                sheet.cell(row, col).value = None
    else:
        book = writer.book

    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    sheet = book[sheet_name]

    if adjust_col_width:
        for ind in range(len(dataframe.columns)):
            max_len = max(len(dataframe.columns[ind]), dataframe.iloc[:, ind].astype(str).map(len).max()) + 1
            sheet.column_dimensions[sheet.cell(1, ind + 1).column_letter].width = max_len

    column_formats = {'money_fmt': '$#,##0.00', 'perc_fmt_no_decimal': '0%',
                             'perc_fmt_decimal': '0.00%', 'decimal_fmt': '0.00'}

    for fmt in column_formats:
        for col in [dataframe.columns.get_loc(c)+1 for c in columns_to_format.get(fmt, [])]:
            for row in range(2, dataframe.shape[0] + 2):
                sheet.cell(row, col).number_format = perc_fmt1

    writer.save(), writer.close()